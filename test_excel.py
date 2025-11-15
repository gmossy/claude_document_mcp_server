#!/usr/bin/env python3
"""
Test Excel support in document_parsers module.
"""

from pathlib import Path
from document_parsers import (
    extract_text_from_excel,
    create_excel_from_data,
    extract_text_from_file
)

def test_excel_support():
    """Test Excel extraction and creation."""
    
    print("🧪 Testing Excel Support")
    print("=" * 50)
    
    # Create test data
    test_data = [
        ["Name", "Age", "Department", "Salary"],
        ["Alice Johnson", "28", "Engineering", "$95,000"],
        ["Bob Smith", "35", "Marketing", "$75,000"],
        ["Carol Davis", "42", "Sales", "$85,000"],
        ["David Wilson", "31", "Engineering", "$90,000"],
        ["Eve Martinez", "29", "HR", "$70,000"],
    ]
    
    # Test 1: Create Excel file
    print("\n1. Creating test Excel file...")
    test_file = Path("test_excel_output.xlsx")
    try:
        create_excel_from_data(
            data=test_data,
            output_path=test_file,
            sheet_name="Employees",
            title="Employee Data"
        )
        print(f"✅ Created: {test_file} ({test_file.stat().st_size} bytes)")
    except Exception as e:
        print(f"❌ Failed to create Excel: {e}")
        return False
    
    # Test 2: Extract text from Excel
    print("\n2. Extracting text from Excel file...")
    try:
        text, metadata = extract_text_from_excel(test_file)
        print(f"✅ Extracted {len(text)} characters")
        print(f"\nMetadata:")
        for key, value in metadata.items():
            print(f"  - {key}: {value}")
        
        print(f"\nExtracted content (first 300 chars):")
        print("-" * 50)
        print(text[:300])
        print("-" * 50)
    except Exception as e:
        print(f"❌ Failed to extract text: {e}")
        return False
    
    # Test 3: Test auto-detection
    print("\n3. Testing auto-detection with extract_text_from_file...")
    try:
        text2, metadata2 = extract_text_from_file(test_file)
        print(f"✅ Auto-detected format: {metadata2.get('format')}")
        print(f"✅ Extracted {len(text2)} characters")
    except Exception as e:
        print(f"❌ Failed auto-detection: {e}")
        return False
    
    # Test 4: Create multi-sheet Excel
    print("\n4. Creating multi-sheet Excel file...")
    from openpyxl import Workbook
    
    multi_file = Path("test_excel_multisheet.xlsx")
    try:
        wb = Workbook()
        
        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.append(["Total Employees", "5"])
        ws1.append(["Total Departments", "4"])
        ws1.append(["Average Salary", "$83,000"])
        
        # Sheet 2: Details
        ws2 = wb.create_sheet("Details")
        for row in test_data:
            ws2.append(row)
        
        # Sheet 3: Notes
        ws3 = wb.create_sheet("Notes")
        ws3.append(["Note", "Description"])
        ws3.append(["1", "All salaries are approximate"])
        ws3.append(["2", "Data as of 2024"])
        
        wb.save(str(multi_file))
        wb.close()
        print(f"✅ Created multi-sheet file: {multi_file}")
        
        # Extract from multi-sheet
        text3, metadata3 = extract_text_from_excel(multi_file)
        print(f"✅ Extracted from {metadata3['sheet_count']} sheets")
        print(f"   Sheet names: {', '.join(metadata3['sheet_names'])}")
        print(f"   Total rows: {metadata3['total_rows']}")
        print(f"   Total cells: {metadata3['total_cells']}")
        
    except Exception as e:
        print(f"❌ Failed multi-sheet test: {e}")
        return False
    
    print("\n" + "=" * 50)
    print("✅ All Excel tests passed!")
    print("\nTest files created:")
    print(f"  - {test_file}")
    print(f"  - {multi_file}")
    
    return True

if __name__ == "__main__":
    import sys
    success = test_excel_support()
    sys.exit(0 if success else 1)
